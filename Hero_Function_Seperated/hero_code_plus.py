
import streamlit as st
import pandas as pd
import re
from io import BytesIO
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import requests
import zipfile
import os
from PIL import Image
from rembg import remove
import io

def clean_sheet_name(name):
    return re.sub(r'[\[\]\*:/\\?]', '', str(name)).strip()[:31]

def parse_hub_blocks(header_row_0, header_row_1, hubs):
    hub_blocks = []
    for idx, cell_hub in enumerate(header_row_0):
        clean_hub = cell_hub.strip()
        if clean_hub in hubs:
            pid1_idx = pid2_idx = None
            for offset in range(4):
                col = idx + offset
                if col >= len(header_row_1):
                    continue
                label = header_row_1[col].strip().lower()
                if label in ("pid1", "pid"):
                    pid1_idx = col
                elif label == "pid2":
                    pid2_idx = col
            hub_blocks.append({
                "Hub": clean_hub,
                "PID1": pid1_idx,
                "PID2": pid2_idx,
            })
    return hub_blocks

def load_mb_images_df(file_or_url):
    if file_or_url is None:
        return None
    try:
        if isinstance(file_or_url, str):
            df = pd.read_csv(file_or_url)
        else:
            df = pd.read_csv(file_or_url)
    except Exception:
        try:
            if isinstance(file_or_url, str):
                df = pd.read_excel(file_or_url)
            else:
                df = pd.read_excel(file_or_url)
        except Exception:
            return None
    if not set(['MB_id', 'image_src']).issubset(df.columns):
        return None
    mapping = {}
    for _, row in df.iterrows():
        mbid = row['MB_id']
        imgsrc = str(row['image_src']).strip()
        if pd.isna(mbid) or pd.isna(imgsrc) or imgsrc == '':
            continue
        try:
            mbid_int = int(str(mbid).strip())
        except Exception:
            continue
        mapping[mbid_int] = {
            "url": f"https://file.milkbasket.com/products/{imgsrc}",
            "filename": os.path.splitext(os.path.basename(imgsrc))[0] + ".png"
        }
    return mapping

def process_csv_campaign_tabs(source_url_or_file, img_pid_map=None):
    if isinstance(source_url_or_file, str):
        df = pd.read_csv(source_url_or_file, header=None)
    else:
        df = pd.read_csv(source_url_or_file, header=None)
    hubs = ['NCR', 'JPR', 'AHM', 'IND', 'MUM', 'Pune', 'BLR', 'HYD', 'CHN', 'SS']
    header_row_0 = df.iloc[0].astype(str).str.strip()
    header_row_1 = df.iloc[1].astype(str).str.strip()
    hub_blocks = parse_hub_blocks(header_row_0, header_row_1, hubs)
    data_rows = df[2:].reset_index(drop=True)
    data_rows[[0, 1, 2]] = data_rows[[0, 1, 2]].fillna(method='ffill')
    asset_index = 2
    focus_col_index = campaign_index = None
    for col in range(len(header_row_1)):
        label = header_row_1[col].strip().lower()
        if label == "focus category/grid":
            focus_col_index = col
        if label == "campaign name":
            campaign_index = col
    combined_tabs = {}
    all_pids_set = set()
    prev_campaign = None
    for idx, row in data_rows.iterrows():
        asset_detail = str(row[asset_index]).strip() if pd.notna(row[asset_index]) else "General"
        campaign_name = str(row[campaign_index]).strip() if campaign_index is not None and pd.notna(row[campaign_index]) else ""
        if campaign_name == "" and prev_campaign is not None:
            campaign_name = prev_campaign
        else:
            prev_campaign = campaign_name
        if asset_detail.lower() == "atc":
            continue
        tab_name = clean_sheet_name(f"{asset_detail}_{campaign_name}")
        for block in hub_blocks:
            hub = block["Hub"]
            pid1 = row[block["PID1"]] if block["PID1"] is not None and pd.notna(row[block["PID1"]]) else ""
            pid2 = row[block["PID2"]] if block["PID2"] is not None and pd.notna(row[block["PID2"]]) else ""
            pid1 = int(pid1) if str(pid1).strip().isdigit() else ""
            pid2 = int(pid2) if str(pid2).strip().isdigit() else ""
            if not pid1 and not pid2:
                continue
            focus_value = str(row[focus_col_index]).strip() if focus_col_index is not None and pd.notna(row[focus_col_index]) else ""
            if pid1: all_pids_set.add(pid1)
            if pid2: all_pids_set.add(pid2)
            img1_url = img_pid_map.get(pid1, {}).get("url", "") if img_pid_map else ""
            img2_url = img_pid_map.get(pid2, {}).get("url", "") if img_pid_map else ""
            row_dict = {
                "Hub": hub,
                "Focus Category/Grid": focus_value,
                "Asset Detail": asset_detail,
                "PID1": pid1,
                "PID2": pid2,
                "Img1": img1_url,
                "Img2": img2_url
            }
            if tab_name not in combined_tabs:
                combined_tabs[tab_name] = []
            combined_tabs[tab_name].append(row_dict)
    for tab_name in combined_tabs:
        focus_set = sorted(set(row["Focus Category/Grid"] for row in combined_tabs[tab_name]))
        def fixed_sort(row):
            hub_index = hubs.index(row["Hub"]) if row["Hub"] in hubs else 999
            cat_index = focus_set.index(row["Focus Category/Grid"]) if row["Focus Category/Grid"] in focus_set else 999
            return (hub_index, cat_index)
        combined_tabs[tab_name] = sorted(combined_tabs[tab_name], key=fixed_sort)
    all_pids_list = sorted(all_pids_set)
    all_pids_tab = pd.DataFrame({
        'PID': all_pids_list,
        'Img': [img_pid_map.get(pid, {}).get("url", "") if img_pid_map else "" for pid in all_pids_list]
    })
    combined_tabs["All_PIDs"] = all_pids_tab.to_dict(orient="records")
    return combined_tabs

def has_transparency(img_bytes):
    try:
        img = Image.open(io.BytesIO(img_bytes))
        if img.mode in ("RGBA", "LA") or (img.mode == "P" and "transparency" in img.info):
            alpha = img.getchannel("A") if "A" in img.getbands() else None
            if alpha and alpha.getextrema()[0] < 255:
                return True
        return False
    except Exception:
        return False


st.set_page_config(layout="wide")
st.title("🧾 HERO CODE PLUS: PID Formatter + Image Filler")
st.markdown("""
Required Columns: ** 
'Category | Campaign Name | Asset Detail | Focus Category/Grid | PID1 | Name1 | PID2 | Name2'

""", unsafe_allow_html=True)

tab = st.radio("1️⃣ Choose campaign file input method:", ["Upload CSV file", "Paste Google Sheet link"])
main_df = None

if tab == "Upload CSV file":
    uploaded_file = st.file_uploader("📤 Upload campaign CSV File", type=["csv"])
    if uploaded_file:
        main_df = uploaded_file
elif tab == "Paste Google Sheet link":
    sheet_url = st.text_input("🔗 Paste your Google Sheet URL")
    if sheet_url:
        match = re.search(r'/d/([a-zA-Z0-9-_]+)', sheet_url)
        if match:
            sheet_id = match.group(1)
            csv_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"
            embed_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/preview"
            st.markdown("### 👁️ Google Sheet Preview")
            st.components.v1.iframe(embed_url, height=500)
            main_df = csv_url
        else:
            st.warning("⚠️ Invalid Google Sheet URL format.")

img_pid_map = None
st.markdown("---")
st.markdown("#### 2️⃣ (Optional) Upload product dump file (with MB_id & image_src), or Google Sheet link")
prod_tab = st.radio("Product dump input method:", ["None", "Upload CSV/Excel file", "Paste Google Sheet link"], index=0)
prod_file_or_url = None
if prod_tab == "Upload CSV/Excel file":
    prod_uploaded = st.file_uploader("📤 Upload product dump CSV or Excel", type=["csv", "xlsx"])
    if prod_uploaded:
        prod_file_or_url = prod_uploaded
elif prod_tab == "Paste Google Sheet link":
    prod_sheet_url = st.text_input("Paste product dump Google Sheet URL")
    if prod_sheet_url:
        match = re.search(r'/d/([a-zA-Z0-9-_]+)', prod_sheet_url)
        if match:
            prod_sheet_id = match.group(1)
            prod_csv_url = f"https://docs.google.com/spreadsheets/d/{prod_sheet_id}/export?format=csv"
            prod_file_or_url = prod_csv_url
        else:
            st.warning("⚠️ Invalid Google Sheet URL format for product dump.")

if prod_file_or_url is not None:
    img_pid_map = load_mb_images_df(prod_file_or_url)

tabs = None
if main_df:
    tabs = process_csv_campaign_tabs(main_df, img_pid_map)
    


if tabs:
    tab_names = list(tabs.keys())
    selected_tab = st.selectbox("👁️ Preview a tab before export:", tab_names)
    preview_df = pd.DataFrame(tabs[selected_tab])
    st.dataframe(preview_df)

    # Excel Export (unchanged)
    output = BytesIO()
    wb = Workbook()
    for tname in tab_names:
        ws = wb.create_sheet(title=clean_sheet_name(tname))
        df_to_write = pd.DataFrame(tabs[tname])
        yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        red_fill = PatternFill(start_color="FF8888", end_color="FF8888", fill_type="solid")
        bold_font = Font(bold=True)
        for r_idx, row in enumerate(dataframe_to_rows(df_to_write, index=False, header=True), 1):
            ws.append(row)
            if r_idx == 1:
                for c_idx in range(1, len(row) + 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.fill = yellow_fill
                    cell.font = bold_font
            if tname == "All_PIDs" and r_idx > 1:
                img_cell = ws.cell(row=r_idx, column=2)
                img_url = img_cell.value
                if not img_url:
                    img_cell.fill = red_fill
                else:
                    try:
                        resp = requests.head(img_url, timeout=4)
                        if resp.status_code != 200:
                            img_cell.fill = red_fill
                    except Exception:
                        img_cell.fill = red_fill
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(output)
    output.seek(0)
    st.success("✅ Done! Your Excel file is ready for download.")
    st.download_button(
        "📥 Download Final Excel",
        output,
        file_name=f"Hero_Code_Plus_Output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # ---------------------
    # THREE STEP IMAGE FLOW
    # ---------------------
    if selected_tab == "All_PIDs" and not preview_df.empty:
        if "image_bytes_list" not in st.session_state:
            st.session_state.image_bytes_list = None

        # 1. Collect Images
        if st.button("Collect Images to Download"):
            collected_images = []
            progress = st.progress(0)
            total = len(preview_df)
            for idx, row in preview_df.iterrows():
                pid = row.get("PID", "")
                img_info = img_pid_map.get(int(pid), {}) if img_pid_map else {}
                img_url = img_info.get("url", "")
                orig_filename = img_info.get("filename", "")
                if img_url and orig_filename:
                    try:
                        r = requests.get(img_url, timeout=10)
                        if r.status_code == 200:
                            collected_images.append({
                                "pid": pid,
                                "img_bytes": r.content,
                                "orig_filename": orig_filename
                            })
                    except Exception:
                        continue
                progress.progress((idx + 1) / total)
            st.session_state.image_bytes_list = collected_images
            st.success(f"Collected {len(collected_images)} images! You can now download.")

        # 2. Download Images.zip (all images as PNG, resized, no rembg)
        if st.session_state.image_bytes_list:
            if st.button("Download Images.zip (resized .png, no BG removal)"):
                MAX_SIZE = (650, 650)
                progress = st.progress(0)
                total = len(st.session_state.image_bytes_list)
                processed = 0
                zip_buffer = BytesIO()
                with st.spinner("Creating ZIP..."):
                    with zipfile.ZipFile(zip_buffer, "w") as zipf:
                        for idx, item in enumerate(st.session_state.image_bytes_list):
                            orig_filename = item["orig_filename"]
                            img_bytes = item["img_bytes"]
                            try:
                                img = Image.open(io.BytesIO(img_bytes)).convert("RGBA")
                                img.thumbnail((MAX_SIZE), Image.Resampling.LANCZOS)
                                img_byte_arr = io.BytesIO()
                                img.save(img_byte_arr, format='PNG')
                                zipf.writestr(orig_filename, img_byte_arr.getvalue())
                                processed += 1
                            except Exception:
                                continue
                            progress.progress((idx + 1) / total)
                zip_buffer.seek(0)
                st.success(f"{processed} images in ZIP (resized, no background removal).")
                st.download_button(
                    label="Download Images.zip",
                    data=zip_buffer,
                    file_name="All_PID_Images.zip",
                    mime="application/zip"
                )

            # 3. Download Rembg Images.zip (only after click)
            if st.button("Rembg Download Images (background removed .png)"):
                MAX_SIZE = (650, 650)
                progress = st.progress(0)
                total = len(st.session_state.image_bytes_list)
                processed = 0
                zip_buffer = BytesIO()
                with st.spinner("Running background removal..."):
                    with zipfile.ZipFile(zip_buffer, "w") as zipf:
                        for idx, item in enumerate(st.session_state.image_bytes_list):
                            orig_filename = item["orig_filename"]
                            img_bytes = item["img_bytes"]
                            try:
                                if has_transparency(img_bytes):
                                    img = Image.open(io.BytesIO(img_bytes)).convert("RGBA")
                                    img.thumbnail((MAX_SIZE), Image.Resampling.LANCZOS)
                                    img_byte_arr = io.BytesIO()
                                    img.save(img_byte_arr, format='PNG')
                                    zipf.writestr(orig_filename, img_byte_arr.getvalue())
                                else:
                                    input_image = Image.open(io.BytesIO(img_bytes)).convert("RGBA")
                                    input_image.thumbnail((MAX_SIZE), Image.Resampling.LANCZOS)
                                    output_image = remove(input_image)
                                    img_byte_arr = io.BytesIO()
                                    output_image.save(img_byte_arr, format='PNG')
                                    zipf.writestr(orig_filename, img_byte_arr.getvalue())
                                processed += 1
                            except Exception:
                                continue
                            progress.progress((idx + 1) / total)
                zip_buffer.seek(0)
                st.success(f"{processed} images (background removed as needed). Download below.")
                st.download_button(
                    label="Rembg Download Images.zip",
                    data=zip_buffer,
                    file_name="All_PID_Images_no_bg.zip",
                    mime="application/zip"
                )
        elif st.session_state.image_bytes_list is not None:
            st.warning("No images were collected. Please check your input data.")
